# ===================== 注会题目PDF转Excel 网页版工具 =====================
# 运行命令：streamlit run 注会题目网页工具.py
# =============================================================================

import streamlit as st
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re
from io import BytesIO

# 全局配置
MAX_SUB_QUESTION = 15  # 最大支持15个小问
TABLE_COLUMNS = [
    "题型", "题目", "分值", "A", "B", "C", "D", "答案", "解析", "资料"
] + [f"要求{i}" for i in range(1, MAX_SUB_QUESTION + 1)] + ["参考答案"]

# 题型匹配规则
QUESTION_TYPE_PATTERNS = {
    "单选题": re.compile(r"^单选题\s*$|^一、单项选择题\s*$"),
    "多选题": re.compile(r"^多选题\s*$|^二、多项选择题\s*$"),
    "计算分析题": re.compile(r"^计算分析题\s*$|^三、计算分析题\s*$"),
    "综合题": re.compile(r"^综合题\s*$|^四、综合题\s*$")
}
# 题目/选项/小问/答案匹配规则
QUESTION_NUM_PATTERN = re.compile(r"^(\d+)\.\s*|^\((\d+)\)\s*")
OPTION_PATTERN = re.compile(r"^([A-D])\.\s*(.*)$")
SUB_QUESTION_PATTERN = re.compile(r"^\((\d+)\)\s*|^(\d+)\.\s*")
ANSWER_PATTERN = re.compile(r"^答案\s*[:：]\s*(.*)$")
ANALYSIS_PATTERN = re.compile(r"^解析\s*[:：]\s*(.*)$")
SCORE_PATTERN = re.compile(r"^本题\s*(\d+)\s*分$|^分值\s*[:：]\s*(\d+)\s*分$")

# 样式定义
HEADER_STYLE = {
    "font": Font(name="微软雅黑", size=11, bold=True, color="FFFFFF"),
    "fill": PatternFill("solid", fgColor="4472C4"),
    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "border": Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
}
DATA_ROW_STYLE = {
    "font": Font(name="微软雅黑", size=10),
    "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
    "border": Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
}
CENTER_CELL_STYLE = {
    "font": Font(name="微软雅黑", size=10),
    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "border": Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
}

# 提取PDF文本
def extract_pdf_text(pdf_file):
    full_text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"
    lines = [line.strip() for line in full_text.split("\n") if line.strip()]
    return lines

# 解析题目
def parse_questions(lines):
    questions = []
    current_question = None
    current_question_type = None
    current_sub_questions = []
    current_answer = ""
    current_analysis = ""
    current_score = 2

    for line in lines:
        # 识别题型
        type_matched = False
        for q_type, pattern in QUESTION_TYPE_PATTERNS.items():
            if pattern.match(line):
                current_question_type = q_type
                type_matched = True
                if q_type in ["计算分析题", "综合题"]:
                    current_score = 9
                break
        if type_matched:
            continue

        # 处理单选/多选题
        if current_question_type in ["单选题", "多选题"]:
            num_match = QUESTION_NUM_PATTERN.match(line)
            if num_match:
                if current_question:
                    req_empty = {f"要求{i}": "" for i in range(1, MAX_SUB_QUESTION + 1)}
                    questions.append({
                        "题型": current_question_type,
                        "题目": current_question,
                        "分值": current_score,
                        "A": current_options.get("A", ""),
                        "B": current_options.get("B", ""),
                        "C": current_options.get("C", ""),
                        "D": current_options.get("D", ""),
                        "答案": current_answer,
                        "解析": current_analysis,
                        "资料": "",
                        **req_empty,
                        "参考答案": current_answer
                    })
                current_question = line[num_match.end():].strip()
                current_options = {}
                current_answer = ""
                current_analysis = ""
                current_score = 2
                continue

            option_match = OPTION_PATTERN.match(line)
            if option_match:
                current_options[option_match.group(1)] = option_match.group(2)
                continue

            answer_match = ANSWER_PATTERN.match(line)
            if answer_match:
                current_answer = answer_match.group(1).strip()
                continue

            analysis_match = ANALYSIS_PATTERN.match(line)
            if analysis_match:
                current_analysis = analysis_match.group(1).strip()
                continue

            score_match = SCORE_PATTERN.match(line)
            if score_match:
                current_score = int(score_match.group(1) if score_match.group(1) else score_match.group(2))
                continue

            if current_question and not any([option_match, answer_match, analysis_match, score_match]):
                current_question += " " + line
                continue

        # 处理计算分析/综合题
        if current_question_type in ["计算分析题", "综合题"]:
            num_match = QUESTION_NUM_PATTERN.match(line)
            if num_match and not current_question:
                current_question = line[num_match.end():].strip()
                current_sub_questions = []
                current_answer = ""
                current_analysis = ""
                current_score = 9
                continue

            sub_q_match = SUB_QUESTION_PATTERN.match(line)
            if sub_q_match:
                sub_q_num = int(sub_q_match.group(1) if sub_q_match.group(1) else sub_q_match.group(2))
                if 1 <= sub_q_num <= MAX_SUB_QUESTION:
                    current_sub_questions.append({
                        "序号": sub_q_num,
                        "内容": line[sub_q_match.end():].strip()
                    })
                continue

            answer_match = ANSWER_PATTERN.match(line)
            if answer_match:
                current_answer = answer_match.group(1).strip()
                continue

            analysis_match = ANALYSIS_PATTERN.match(line)
            if analysis_match:
                current_analysis = analysis_match.group(1).strip()
                continue

            score_match = SCORE_PATTERN.match(line)
            if score_match:
                current_score = int(score_match.group(1) if score_match.group(1) else score_match.group(2))
                continue

            if current_question and not any([sub_q_match, answer_match, analysis_match, score_match]):
                current_question += " " + line
                continue

            if line.startswith("本题结束") or line.startswith("参考答案") or (current_question and len(current_sub_questions) > 0 and not line.strip()):
                req_dict = {f"要求{i}": "" for i in range(1, MAX_SUB_QUESTION + 1)}
                for sub_q in current_sub_questions:
                    req_dict[f"要求{sub_q['序号']}"] = sub_q['内容']
                questions.append({
                    "题型": current_question_type,
                    "题目": current_question,
                    "分值": current_score,
                    "A": "", "B": "", "C": "", "D": "",
                    "答案": "",
                    "解析": current_analysis,
                    "资料": "",
                    **req_dict,
                    "参考答案": current_answer
                })
                current_question = None
                current_sub_questions = []
                continue

    # 保存最后一题
    if current_question:
        if current_question_type in ["单选题", "多选题"]:
            req_empty = {f"要求{i}": "" for i in range(1, MAX_SUB_QUESTION + 1)}
            questions.append({
                "题型": current_question_type,
                "题目": current_question,
                "分值": current_score,
                "A": current_options.get("A", ""),
                "B": current_options.get("B", ""),
                "C": current_options.get("C", ""),
                "D": current_options.get("D", ""),
                "答案": current_answer,
                "解析": current_analysis,
                "资料": "",
                **req_empty,
                "参考答案": current_answer
            })
        elif current_question_type in ["计算分析题", "综合题"]:
            req_dict = {f"要求{i}": "" for i in range(1, MAX_SUB_QUESTION + 1)}
            for sub_q in current_sub_questions:
                req_dict[f"要求{sub_q['序号']}"] = sub_q['内容']
            questions.append({
                "题型": current_question_type,
                "题目": current_question,
                "分值": current_score,
                "A": "", "B": "", "C": "", "D": "",
                "答案": "",
                "解析": current_analysis,
                "资料": "",
                **req_dict,
                "参考答案": current_answer
            })

    return questions

# 生成Excel文件
def generate_excel(questions):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "注会题目"

    # 写入表头
    for col_idx, col_name in enumerate(TABLE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_STYLE["font"]
        cell.fill = HEADER_STYLE["fill"]
        cell.alignment = HEADER_STYLE["alignment"]
        cell.border = HEADER_STYLE["border"]

    # 写入数据
    for row_idx, question in enumerate(questions, 2):
        for col_idx, col_name in enumerate(TABLE_COLUMNS, 1):
            cell_value = question.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if col_name in ["题型", "分值", "答案"]:
                cell.font = CENTER_CELL_STYLE["font"]
                cell.alignment = CENTER_CELL_STYLE["alignment"]
                cell.border = CENTER_CELL_STYLE["border"]
            else:
                cell.font = DATA_ROW_STYLE["font"]
                cell.alignment = DATA_ROW_STYLE["alignment"]
                cell.border = DATA_ROW_STYLE["border"]

    # 自动调整列宽
    for col_idx in range(1, len(TABLE_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = max(min(max_length + 2, 50), 10)
        ws.column_dimensions[col_letter].width = adjusted_width

    # 冻结表头
    ws.freeze_panes = "A2"

    # 设置行高
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20

    # 保存到内存
    wb.save(output)
    output.seek(0)
    return output, len(questions)

# 网页主界面
def main():
    st.set_page_config(page_title="注会题目PDF转Excel工具", page_icon="📊", layout="wide")
    st.title("📊 注会题目PDF转Excel 一键生成工具")
    st.markdown("支持单选/多选/计算分析/综合题，自动拆分选项、小问、答案、解析，生成规范Excel表格")

    # 上传PDF
    uploaded_file = st.file_uploader("📤 上传注会题目PDF文件", type="pdf", help="仅支持PDF格式文件")

    if uploaded_file:
        with st.spinner("正在读取PDF文件..."):
            try:
                # 提取文本
                pdf_lines = extract_pdf_text(uploaded_file)
                st.success(f"✅ PDF读取完成，共提取 {len(pdf_lines)} 行文本")

                # 解析题目
                with st.spinner("正在解析题目内容..."):
                    questions = parse_questions(pdf_lines)
                    st.success(f"✅ 题目解析完成，共识别 {len(questions)} 道题目")

                # 生成Excel
                with st.spinner("正在生成Excel表格..."):
                    excel_file, question_count = generate_excel(questions)
                    st.success(f"✅ Excel生成完成，共生成 {question_count} 道题目")

                # 下载按钮
                st.download_button(
                    label="📥 下载生成的Excel文件",
                    data=excel_file,
                    file_name=f"注会题目_生成结果.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                # 预览题目列表
                st.subheader("📋 解析的题目列表预览")
                for i, q in enumerate(questions, 1):
                    with st.expander(f"{i}. {q['题型']} - {q['题目'][:50]}..."):
                        st.write(f"**分值**：{q['分值']}分")
                        if q['A']:
                            st.write(f"**A选项**：{q['A']}")
                            st.write(f"**B选项**：{q['B']}")
                            st.write(f"**C选项**：{q['C']}")
                            st.write(f"**D选项**：{q['D']}")
                        if q['答案']:
                            st.write(f"**答案**：{q['答案']}")
                        if q['解析']:
                            st.write(f"**解析**：{q['解析']}")
                        for j in range(1, MAX_SUB_QUESTION + 1):
                            req_key = f"要求{j}"
                            if q[req_key]:
                                st.write(f"**要求{j}**：{q[req_key]}")

            except Exception as e:
                st.error(f"❌ 处理出错：{str(e)}")
                st.markdown("请检查您的PDF文件是否正常，或联系作者优化代码")

    # 页脚说明
    st.markdown("---")
    st.markdown("💡 工具说明：")
    st.markdown("1. 支持注会考试真题的标准格式，自动识别题型、选项、小问、答案、解析")
    st.markdown("2. 计算分析/综合题最多支持15个小问，自动对应到要求1-15列，无对应小问自动留空")
    st.markdown("3. 生成的Excel表格带专业样式，自动换行、列宽适配、冻结表头，可直接打印或导入题库")

if __name__ == "__main__":
    main()
